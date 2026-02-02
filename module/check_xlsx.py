from module.log import Log
from openpyxl import load_workbook

# 定义一个logger
logging = Log(__name__).getlog()

def check_data_types(file_path, expected_types):
    
    error_num = 0
    
    try:
        # 读取Excel文件
        workbook = load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames
        logging.debug(f"Sheet names in the workbook: {sheet_names}")
    except Exception as e:
        logging.error(f"Error reading the Excel file: {e}")
        return

    for sheet_name, columns in expected_types.items():

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            logging.info(f"Checking sheet: {sheet_name}")

            # 获取标题行
            headers = {cell.value: cell.column for cell in sheet[1]}

            for column, expected_type in columns.items():
                if column in headers:
                    col_index = headers[column]

                    for row in range(2, sheet.max_row + 1):
                        cell_value = sheet.cell(row=row, column=col_index).value

                        if cell_value is not None and isinstance(cell_value, expected_type):
                            logging.debug(f"Column {column} in {sheet_name} is of expected type {expected_type}")
                            if '时间' in column:
                                datatime = int(cell_value) - 946656000
                                if datatime % 86400 != 0:
                                    logging.error(f"Invalid datetime in {sheet_name} - {column} at row {row}: {cell_value}")
                                    error_num += 1
                        
                        else:
                            logging.error(f"Type mismatch in {sheet_name} - {column}: expected {expected_type}, but got {type(cell_value)} at row {row}")
                            error_num += 1
                            
                else:
                    logging.error(f"Column {column} not found in {sheet_name}")
                    error_num += 1
        else:
            logging.error(f"No sheet named {sheet_name} in the workbook")
            error_num += 1
            
    return error_num



